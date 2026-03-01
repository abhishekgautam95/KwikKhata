import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from scripts.migrate_excel import migrate_excel


class TestMigration(unittest.TestCase):
    def setUp(self):
        fd, self.path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        os.remove(self.path)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Khata'
        ws.append(['Customer_Name', 'Balance'])
        ws.append(['Raju', 250])
        wb.save(self.path)
        wb.close()

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_add_transactions_sheet_and_backfill(self):
        created, added = migrate_excel(self.path, backfill_opening_entries=True)
        self.assertGreaterEqual(created, 1)
        self.assertEqual(added, 1)

        wb = load_workbook(self.path)
        try:
            self.assertIn('Transactions', wb.sheetnames)
            tx = wb['Transactions']
            self.assertEqual(tx.max_row, 2)
        finally:
            wb.close()


if __name__ == '__main__':
    unittest.main()
