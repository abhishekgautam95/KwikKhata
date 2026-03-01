"""
Migration utility for legacy KwikKhata Excel files.

What it does:
1) Ensures `Khata` and `Transactions` sheets exist.
2) Adds required header rows if missing.
3) Optionally backfills a synthetic opening transaction per customer
   from existing non-zero balances when transaction history is empty.
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook

LEDGER_SHEET = "Khata"
HISTORY_SHEET = "Transactions"
HEADERS = ["Customer_Name", "Balance"]
HISTORY_HEADERS = ["Timestamp", "Customer_Name", "Amount", "New_Balance"]


def ensure_headers(ws, expected_headers: list[str]) -> None:
    first_row = [ws.cell(row=1, column=i + 1).value for i in range(len(expected_headers))]
    if first_row != expected_headers:
        ws.delete_rows(1, 1)
        ws.insert_rows(1)
        for idx, header in enumerate(expected_headers, start=1):
            ws.cell(row=1, column=idx, value=header)


def migrate_excel(path: str, backfill_opening_entries: bool = True) -> tuple[int, int]:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    try:
        wb = load_workbook(file_path)
    except BadZipFile as exc:
        raise RuntimeError(
            "Excel file is corrupted (BadZipFile). Restore from backup and retry migration."
        ) from exc
    created_sheets = 0
    added_entries = 0
    try:
        if LEDGER_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(LEDGER_SHEET, 0)
            ws.append(HEADERS)
            created_sheets += 1

        if HISTORY_SHEET not in wb.sheetnames:
            hs = wb.create_sheet(HISTORY_SHEET)
            hs.append(HISTORY_HEADERS)
            created_sheets += 1

        ledger = wb[LEDGER_SHEET]
        history = wb[HISTORY_SHEET]

        ensure_headers(ledger, HEADERS)
        ensure_headers(history, HISTORY_HEADERS)

        history_has_rows = history.max_row > 1
        if backfill_opening_entries and not history_has_rows:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for row in range(2, ledger.max_row + 1):
                name = ledger.cell(row=row, column=1).value
                balance = ledger.cell(row=row, column=2).value or 0
                try:
                    balance = float(balance)
                except (TypeError, ValueError):
                    continue
                if name and balance != 0:
                    history.append([ts, str(name), balance, balance])
                    added_entries += 1

        wb.save(file_path)
        return created_sheets, added_entries
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Migrate KwikKhata Excel format.")
    parser.add_argument("--file", default="kwik_khata_db.xlsx", help="Path to Excel DB")
    parser.add_argument(
        "--no-backfill",
        action="store_true",
        help="Do not create synthetic opening entries in Transactions",
    )
    args = parser.parse_args()

    try:
        created_sheets, added_entries = migrate_excel(
            path=args.file,
            backfill_opening_entries=not args.no_backfill,
        )
        print(
            f"Migration complete: created_sheets={created_sheets}, added_opening_entries={added_entries}"
        )
    except Exception as exc:
        print(f"Migration failed: {exc}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
