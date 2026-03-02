"""
database.py — KwikKhata Ledger Database Module
===============================================
Excel-backed ledger store using openpyxl.
"""

from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
import json

from openpyxl import Workbook, load_workbook

try:
    import psycopg
except Exception:  # pragma: no cover - optional dependency for postgres backend
    psycopg = None


class KhataDB:
    """Excel-backed customer ledger database."""

    DB_FILE = "kwik_khata_db.xlsx"
    LEDGER_SHEET = "Khata"
    HISTORY_SHEET = "Transactions"
    HEADERS = ["Customer_Name", "Balance"]
    HISTORY_HEADERS = ["Timestamp", "Customer_Name", "Amount", "New_Balance"]
    BACKUP_DIR = "backups"
    NAME_NOISE_WORDS = {
        "rupees",
        "rupaye",
        "rupay",
        "rs",
        "inr",
        "entry",
        "entries",
        "pending",
        "show",
        "dikha",
        "dikhao",
        "check",
        "please",
        "pls",
        "chini",
        "cheeni",
        "daal",
        "dal",
        "namkeen",
        "biscuit",
        "biskut",
        "sabun",
        "soap",
        "tel",
        "oil",
        "atta",
        "chawal",
        "rice",
        "salt",
        "namak",
        "masala",
        "mr",
        "mrs",
        "miss",
        "shri",
    }
    NAME_STOPWORDS = {"ka", "ki", "ko", "ne", "se", "mein", "me", "hai", "tha", "thi"}
    NAME_TITLE_SUFFIX = {"ji", "bhai", "bhaiya", "sir", "madam", "uncle", "aunty"}

    def __init__(self, filepath: str | None = None):
        self.filepath = filepath or self.DB_FILE
        self.backup_dir = os.getenv("KWIKKHATA_BACKUP_DIR", self.BACKUP_DIR)
        self.backup_keep = int(os.getenv("KWIKKHATA_BACKUP_KEEP", "20"))
        self.enable_backups = os.getenv("KWIKKHATA_ENABLE_BACKUP", "true").strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        self._ensure_file_exists()
        self._backup_if_enabled()

    def _ensure_file_exists(self) -> None:
        if not os.path.exists(self.filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = self.LEDGER_SHEET
            ws.append(self.HEADERS)
            history = wb.create_sheet(self.HISTORY_SHEET)
            history.append(self.HISTORY_HEADERS)
            wb.save(self.filepath)
            wb.close()
            return

        # Ensure mandatory sheets exist even if file already present.
        wb = load_workbook(self.filepath)
        try:
            if self.LEDGER_SHEET not in wb.sheetnames:
                ws = wb.create_sheet(self.LEDGER_SHEET, 0)
                ws.append(self.HEADERS)
            if self.HISTORY_SHEET not in wb.sheetnames:
                history = wb.create_sheet(self.HISTORY_SHEET)
                history.append(self.HISTORY_HEADERS)
            wb.save(self.filepath)
        finally:
            wb.close()

    def _normalize_name(self, name: str) -> str:
        return str(name).strip().title()

    def _canonicalize_customer_name(self, raw_name: str) -> str:
        text = re.sub(r"[^a-zA-Z\s]", " ", str(raw_name).lower())
        tokens = [t for t in re.sub(r"\s+", " ", text).strip().split(" ") if t]
        filtered = [
            token
            for token in tokens
            if token not in self.NAME_NOISE_WORDS and token not in self.NAME_STOPWORDS
        ]
        if not filtered:
            return ""
        if len(filtered) == 1:
            return self._normalize_name(filtered[0])
        if filtered[-1] in self.NAME_TITLE_SUFFIX:
            return self._normalize_name(" ".join(filtered[-2:]))
        return self._normalize_name(" ".join(filtered[:2]))

    def _find_customer_row(self, ws, name: str) -> int | None:
        normalized = self._normalize_name(name)
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value and self._normalize_name(value) == normalized:
                return row
        return None

    def _append_history(self, ws_history, name: str, amount: float, new_balance: float) -> None:
        ws_history.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                self._normalize_name(name),
                round(float(amount), 2),
                round(float(new_balance), 2),
            ]
        )

    def _backup_if_enabled(self) -> None:
        """Create a timestamped DB backup and prune older backups."""
        if not self.enable_backups:
            return

        db_path = Path(self.filepath)
        if not db_path.exists():
            return

        backup_dir = Path(self.backup_dir)
        backup_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{db_path.stem}_{timestamp}{db_path.suffix}"
        backup_path = backup_dir / backup_name
        shutil.copy2(db_path, backup_path)

        # Keep only the latest N backups.
        backups = sorted(
            backup_dir.glob(f"{db_path.stem}_*{db_path.suffix}"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in backups[self.backup_keep :]:
            old.unlink(missing_ok=True)

    def add_transaction(self, name: str, amount: int | float) -> float:
        """Add positive (udhaar) or negative (jama) amount to customer balance."""
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            raise ValueError("customer name cannot be empty")

        try:
            amount = round(float(amount), 2)
        except (TypeError, ValueError) as exc:
            raise ValueError("amount must be a number") from exc

        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            ws_history = wb[self.HISTORY_SHEET]
            row = self._find_customer_row(ws, normalized_name)

            if row:
                current_balance = float(ws.cell(row=row, column=2).value or 0)
                new_balance = round(current_balance + amount, 2)
                ws.cell(row=row, column=2, value=new_balance)
            else:
                new_balance = round(amount, 2)
                ws.append([normalized_name, new_balance])

            self._append_history(ws_history, normalized_name, amount, new_balance)
            wb.save(self.filepath)
            return new_balance
        finally:
            wb.close()

    def get_balance(self, name: str) -> float | None:
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            return None

        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            row = self._find_customer_row(ws, normalized_name)
            if not row:
                return None
            return round(float(ws.cell(row=row, column=2).value or 0), 2)
        finally:
            wb.close()

    def get_all_ledgers(self) -> list[dict]:
        """Return customers with pending positive balances."""
        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            ledgers: list[dict] = []
            for row in range(2, ws.max_row + 1):
                name = ws.cell(row=row, column=1).value
                balance = float(ws.cell(row=row, column=2).value or 0)
                if name and balance > 0:
                    ledgers.append({"name": str(name), "balance": round(balance, 2)})
            return ledgers
        finally:
            wb.close()

    def get_recent_transactions(self, limit: int = 10) -> list[dict]:
        """Return most recent transaction rows for debugging/operator visibility."""
        limit = max(1, int(limit))
        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.HISTORY_SHEET]
            rows = []
            start_row = max(2, ws.max_row - limit + 1)
            for row in range(ws.max_row, start_row - 1, -1):
                rows.append(
                    {
                        "timestamp": ws.cell(row=row, column=1).value,
                        "name": ws.cell(row=row, column=2).value,
                        "amount": ws.cell(row=row, column=3).value,
                        "new_balance": ws.cell(row=row, column=4).value,
                    }
                )
            return rows
        finally:
            wb.close()

    def get_customer_transactions(self, name: str, limit: int = 10) -> list[dict]:
        """Return most recent transactions for a specific customer."""
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            return []

        limit = max(1, int(limit))
        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.HISTORY_SHEET]
            rows: list[dict] = []
            for row in range(ws.max_row, 1, -1):
                row_name = str(ws.cell(row=row, column=2).value or "").strip()
                if self._normalize_name(row_name) != normalized_name:
                    continue
                rows.append(
                    {
                        "timestamp": ws.cell(row=row, column=1).value,
                        "name": ws.cell(row=row, column=2).value,
                        "amount": ws.cell(row=row, column=3).value,
                        "new_balance": ws.cell(row=row, column=4).value,
                    }
                )
                if len(rows) >= limit:
                    break
            return rows
        finally:
            wb.close()

    def undo_last_transaction(self) -> dict | None:
        """Revert most recent transaction entry and return undo summary."""
        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            ws_history = wb[self.HISTORY_SHEET]
            if ws_history.max_row < 2:
                return None

            row = ws_history.max_row
            name = str(ws_history.cell(row=row, column=2).value or "").strip()
            amount = float(ws_history.cell(row=row, column=3).value or 0)
            new_balance = float(ws_history.cell(row=row, column=4).value or 0)
            previous_balance = round(new_balance - amount, 2)

            customer_row = self._find_customer_row(ws, name)
            if customer_row is not None:
                if abs(previous_balance) < 1e-9:
                    ws.delete_rows(customer_row, 1)
                else:
                    ws.cell(row=customer_row, column=2, value=previous_balance)

            ws_history.delete_rows(row, 1)
            wb.save(self.filepath)
            return {
                "customer_name": self._normalize_name(name),
                "amount": round(amount, 2),
                "new_balance": round(previous_balance, 2),
            }
        finally:
            wb.close()

    def get_pending_ledgers_with_age(self) -> list[dict]:
        """
        Return positive-balance ledgers with approximate pending age in days.
        Age is estimated from the last timestamp where running balance crossed
        from <=0 to >0 for that customer.
        """
        wb = load_workbook(self.filepath)
        try:
            ws_ledger = wb[self.LEDGER_SHEET]
            ws_history = wb[self.HISTORY_SHEET]

            current_balances: dict[str, float] = {}
            for row in range(2, ws_ledger.max_row + 1):
                name = ws_ledger.cell(row=row, column=1).value
                balance = float(ws_ledger.cell(row=row, column=2).value or 0)
                if name:
                    current_balances[self._normalize_name(str(name))] = balance

            # replay transactions in chronological order
            state: dict[str, float] = {}
            became_pending_at: dict[str, datetime | None] = {}
            for row in range(2, ws_history.max_row + 1):
                ts_raw = ws_history.cell(row=row, column=1).value
                name_raw = ws_history.cell(row=row, column=2).value
                amount = float(ws_history.cell(row=row, column=3).value or 0)
                if not name_raw:
                    continue
                name = self._normalize_name(str(name_raw))
                prev = state.get(name, 0.0)
                now = round(prev + amount, 2)
                state[name] = now
                if prev <= 0 < now:
                    parsed_ts: datetime | None = None
                    if isinstance(ts_raw, datetime):
                        parsed_ts = ts_raw
                    elif isinstance(ts_raw, str):
                        try:
                            parsed_ts = datetime.strptime(ts_raw, "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            parsed_ts = None
                    became_pending_at[name] = parsed_ts or datetime.now()
                if now <= 0:
                    became_pending_at[name] = None

            now_dt = datetime.now()
            rows: list[dict] = []
            for name, balance in current_balances.items():
                if balance <= 0:
                    continue
                started = became_pending_at.get(name)
                days = 0
                if started is not None:
                    days = max(0, (now_dt - started).days)
                rows.append({"name": name, "balance": round(balance, 2), "pending_days": days})
            return rows
        finally:
            wb.close()

    def merge_customers(self, source_name: str, target_name: str) -> dict:
        source = self._normalize_name(source_name)
        target = self._normalize_name(target_name)
        if not source or not target:
            return {"ok": False, "reason": "source/target name missing"}
        if source == target:
            return {"ok": False, "reason": "source and target are same"}

        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            ws_history = wb[self.HISTORY_SHEET]

            source_row = self._find_customer_row(ws, source)
            if source_row is None:
                return {"ok": False, "reason": f"'{source}' not found"}

            source_balance = float(ws.cell(row=source_row, column=2).value or 0)
            target_row = self._find_customer_row(ws, target)
            if target_row is not None:
                target_balance = float(ws.cell(row=target_row, column=2).value or 0)
                ws.cell(row=target_row, column=2, value=round(target_balance + source_balance, 2))
                ws.delete_rows(source_row, 1)
            else:
                ws.cell(row=source_row, column=1, value=target)

            history_updates = 0
            for row in range(2, ws_history.max_row + 1):
                row_name = self._normalize_name(str(ws_history.cell(row=row, column=2).value or ""))
                if row_name == source:
                    ws_history.cell(row=row, column=2, value=target)
                    history_updates += 1

            target_row_after = self._find_customer_row(ws, target)
            new_balance = 0.0
            if target_row_after is not None:
                new_balance = float(ws.cell(row=target_row_after, column=2).value or 0)
            wb.save(self.filepath)
            return {
                "ok": True,
                "source": source,
                "target": target,
                "history_updates": history_updates,
                "new_balance": round(float(new_balance), 2),
            }
        finally:
            wb.close()

    def cleanup_noisy_customer_names(self) -> dict:
        wb = load_workbook(self.filepath)
        try:
            ws = wb[self.LEDGER_SHEET]
            mappings: dict[str, str] = {}
            for row in range(2, ws.max_row + 1):
                raw_name = ws.cell(row=row, column=1).value
                if not raw_name:
                    continue
                old_name = self._normalize_name(str(raw_name))
                new_name = self._canonicalize_customer_name(old_name)
                if new_name and new_name != old_name:
                    mappings[old_name] = new_name
        finally:
            wb.close()

        if not mappings:
            return {"ok": True, "updated": 0, "details": []}

        details: list[dict] = []
        for old_name, new_name in mappings.items():
            result = self.merge_customers(old_name, new_name)
            if result.get("ok"):
                details.append(result)
        return {"ok": True, "updated": len(details), "details": details}


class PostgresKhataDB:
    """PostgreSQL-backed customer ledger database."""

    NAME_NOISE_WORDS = KhataDB.NAME_NOISE_WORDS
    NAME_STOPWORDS = KhataDB.NAME_STOPWORDS
    NAME_TITLE_SUFFIX = KhataDB.NAME_TITLE_SUFFIX

    def __init__(self, dsn: str | None = None):
        if psycopg is None:
            raise RuntimeError(
                "PostgreSQL backend selected but psycopg is not installed. "
                "Install dependency: pip install psycopg[binary]"
            )
        self.dsn = dsn or os.getenv("KWIKKHATA_DATABASE_URL") or os.getenv("DATABASE_URL", "")
        if not self.dsn:
            raise RuntimeError("PostgreSQL backend selected but KWIKKHATA_DATABASE_URL/DATABASE_URL is missing.")
        self.shop_name = os.getenv("KWIKKHATA_SHOP_NAME", "Default Shop").strip() or "Default Shop"
        self._ensure_schema()

    def _connect(self):
        return psycopg.connect(self.dsn)

    def _normalize_name(self, name: str) -> str:
        return str(name).strip().title()

    def _canonicalize_customer_name(self, raw_name: str) -> str:
        text = re.sub(r"[^a-zA-Z\s]", " ", str(raw_name).lower())
        tokens = [t for t in re.sub(r"\s+", " ", text).strip().split(" ") if t]
        filtered = [
            token
            for token in tokens
            if token not in self.NAME_NOISE_WORDS and token not in self.NAME_STOPWORDS
        ]
        if not filtered:
            return ""
        if len(filtered) == 1:
            return self._normalize_name(filtered[0])
        if filtered[-1] in self.NAME_TITLE_SUFFIX:
            return self._normalize_name(" ".join(filtered[-2:]))
        return self._normalize_name(" ".join(filtered[:2]))

    def _ensure_schema(self) -> None:
        with self._connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS shops (
                        id BIGSERIAL PRIMARY KEY,
                        name TEXT NOT NULL UNIQUE,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    );
                    CREATE TABLE IF NOT EXISTS users (
                        id BIGSERIAL PRIMARY KEY,
                        shop_id BIGINT NOT NULL REFERENCES shops(id) ON DELETE CASCADE,
                        role TEXT NOT NULL DEFAULT 'owner',
                        phone_number TEXT,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    );
                    CREATE TABLE IF NOT EXISTS customers (
                        id BIGSERIAL PRIMARY KEY,
                        shop_id BIGINT NOT NULL REFERENCES shops(id) ON DELETE CASCADE,
                        name TEXT NOT NULL,
                        name_norm TEXT NOT NULL,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (shop_id, name_norm)
                    );
                    CREATE TABLE IF NOT EXISTS ledger_balances (
                        customer_id BIGINT PRIMARY KEY REFERENCES customers(id) ON DELETE CASCADE,
                        balance NUMERIC(14,2) NOT NULL DEFAULT 0,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    );
                    CREATE TABLE IF NOT EXISTS transactions (
                        id BIGSERIAL PRIMARY KEY,
                        shop_id BIGINT NOT NULL REFERENCES shops(id) ON DELETE CASCADE,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        customer_id BIGINT NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                        amount NUMERIC(14,2) NOT NULL,
                        new_balance NUMERIC(14,2) NOT NULL,
                        source TEXT NOT NULL DEFAULT 'system',
                        actor TEXT NOT NULL DEFAULT 'system',
                        idempotency_key TEXT
                    );
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_transactions_shop_idem
                    ON transactions(shop_id, idempotency_key)
                    WHERE idempotency_key IS NOT NULL;
                    CREATE TABLE IF NOT EXISTS idempotency_keys (
                        shop_id BIGINT NOT NULL REFERENCES shops(id) ON DELETE CASCADE,
                        key TEXT NOT NULL,
                        response_json JSONB NOT NULL,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        expires_at TIMESTAMPTZ,
                        PRIMARY KEY (shop_id, key)
                    );
                    CREATE TABLE IF NOT EXISTS audit_logs (
                        id BIGSERIAL PRIMARY KEY,
                        shop_id BIGINT NOT NULL REFERENCES shops(id) ON DELETE CASCADE,
                        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        action TEXT NOT NULL,
                        entity_type TEXT NOT NULL,
                        entity_id TEXT,
                        before_state JSONB,
                        after_state JSONB,
                        meta JSONB,
                        actor TEXT NOT NULL DEFAULT 'system',
                        source TEXT NOT NULL DEFAULT 'system'
                    );
                    """
                )
                cur.execute("INSERT INTO shops(name) VALUES (%s) ON CONFLICT (name) DO NOTHING", (self.shop_name,))

    def _get_shop_id(self, cur) -> int:
        cur.execute("SELECT id FROM shops WHERE name = %s", (self.shop_name,))
        row = cur.fetchone()
        if not row:
            raise RuntimeError("default shop missing in PostgreSQL backend")
        return int(row[0])

    def _get_or_create_customer(self, cur, shop_id: int, normalized_name: str) -> tuple[int, str]:
        cur.execute(
            """
            INSERT INTO customers(shop_id, name, name_norm)
            VALUES (%s, %s, %s)
            ON CONFLICT (shop_id, name_norm)
            DO UPDATE SET name = EXCLUDED.name
            RETURNING id, name
            """,
            (shop_id, normalized_name, normalized_name),
        )
        row = cur.fetchone()
        return int(row[0]), str(row[1])

    def _audit(
        self,
        cur,
        *,
        shop_id: int,
        action: str,
        entity_type: str,
        entity_id: str,
        before_state: dict | None,
        after_state: dict | None,
        meta: dict | None = None,
        actor: str = "system",
        source: str = "system",
    ) -> None:
        cur.execute(
            """
            INSERT INTO audit_logs(
                shop_id, action, entity_type, entity_id, before_state, after_state, meta, actor, source
            )
            VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s, %s)
            """,
            (
                shop_id,
                action,
                entity_type,
                entity_id,
                json.dumps(before_state or {}),
                json.dumps(after_state or {}),
                json.dumps(meta or {}),
                actor,
                source,
            ),
        )

    def add_transaction(
        self,
        name: str,
        amount: int | float,
        *,
        actor: str = "system",
        source: str = "system",
        idempotency_key: str | None = None,
    ) -> float:
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            raise ValueError("customer name cannot be empty")

        try:
            amount = round(float(amount), 2)
        except (TypeError, ValueError) as exc:
            raise ValueError("amount must be a number") from exc

        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                if idempotency_key:
                    cur.execute(
                        """
                        SELECT new_balance
                        FROM transactions
                        WHERE shop_id = %s AND idempotency_key = %s
                        ORDER BY id DESC
                        LIMIT 1
                        """,
                        (shop_id, idempotency_key),
                    )
                    row = cur.fetchone()
                    if row is not None:
                        return round(float(row[0]), 2)

                customer_id, customer_name = self._get_or_create_customer(cur, shop_id, normalized_name)
                cur.execute(
                    "SELECT balance FROM ledger_balances WHERE customer_id = %s FOR UPDATE",
                    (customer_id,),
                )
                bal_row = cur.fetchone()
                current_balance = round(float(bal_row[0]), 2) if bal_row else 0.0
                new_balance = round(current_balance + amount, 2)
                cur.execute(
                    """
                    INSERT INTO ledger_balances(customer_id, balance, updated_at)
                    VALUES (%s, %s, NOW())
                    ON CONFLICT (customer_id)
                    DO UPDATE SET balance = EXCLUDED.balance, updated_at = NOW()
                    """,
                    (customer_id, new_balance),
                )
                cur.execute(
                    """
                    INSERT INTO transactions(shop_id, customer_id, amount, new_balance, actor, source, idempotency_key)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (shop_id, customer_id, amount, new_balance, actor, source, idempotency_key),
                )
                self._audit(
                    cur,
                    shop_id=shop_id,
                    action="add_transaction",
                    entity_type="customer",
                    entity_id=str(customer_id),
                    before_state={"name": customer_name, "balance": current_balance},
                    after_state={"name": customer_name, "balance": new_balance, "amount": amount},
                    actor=actor,
                    source=source,
                )
                return new_balance

    def get_balance(self, name: str) -> float | None:
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            return None
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT lb.balance
                    FROM customers c
                    JOIN ledger_balances lb ON lb.customer_id = c.id
                    WHERE c.shop_id = %s AND c.name_norm = %s
                    """,
                    (shop_id, normalized_name),
                )
                row = cur.fetchone()
                if not row:
                    return None
                return round(float(row[0]), 2)

    def get_all_ledgers(self) -> list[dict]:
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT c.name, lb.balance
                    FROM customers c
                    JOIN ledger_balances lb ON lb.customer_id = c.id
                    WHERE c.shop_id = %s AND lb.balance > 0
                    ORDER BY lb.balance DESC, c.name ASC
                    """,
                    (shop_id,),
                )
                return [{"name": str(r[0]), "balance": round(float(r[1]), 2)} for r in cur.fetchall()]

    def get_recent_transactions(self, limit: int = 10) -> list[dict]:
        limit = max(1, int(limit))
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT t.created_at, c.name, t.amount, t.new_balance
                    FROM transactions t
                    JOIN customers c ON c.id = t.customer_id
                    WHERE t.shop_id = %s
                    ORDER BY t.id DESC
                    LIMIT %s
                    """,
                    (shop_id, limit),
                )
                rows = cur.fetchall()
                return [
                    {
                        "timestamp": r[0].strftime("%Y-%m-%d %H:%M:%S") if hasattr(r[0], "strftime") else str(r[0]),
                        "name": str(r[1]),
                        "amount": round(float(r[2]), 2),
                        "new_balance": round(float(r[3]), 2),
                    }
                    for r in rows
                ]

    def get_customer_transactions(self, name: str, limit: int = 10) -> list[dict]:
        normalized_name = self._normalize_name(name)
        if not normalized_name:
            return []
        limit = max(1, int(limit))
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT t.created_at, c.name, t.amount, t.new_balance
                    FROM transactions t
                    JOIN customers c ON c.id = t.customer_id
                    WHERE t.shop_id = %s AND c.name_norm = %s
                    ORDER BY t.id DESC
                    LIMIT %s
                    """,
                    (shop_id, normalized_name, limit),
                )
                rows = cur.fetchall()
                return [
                    {
                        "timestamp": r[0].strftime("%Y-%m-%d %H:%M:%S") if hasattr(r[0], "strftime") else str(r[0]),
                        "name": str(r[1]),
                        "amount": round(float(r[2]), 2),
                        "new_balance": round(float(r[3]), 2),
                    }
                    for r in rows
                ]

    def undo_last_transaction(self) -> dict | None:
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT t.id, t.customer_id, c.name, t.amount, t.new_balance
                    FROM transactions t
                    JOIN customers c ON c.id = t.customer_id
                    WHERE t.shop_id = %s
                    ORDER BY t.id DESC
                    LIMIT 1
                    FOR UPDATE
                    """,
                    (shop_id,),
                )
                row = cur.fetchone()
                if not row:
                    return None

                tx_id = int(row[0])
                customer_id = int(row[1])
                customer_name = str(row[2])
                amount = round(float(row[3]), 2)
                new_balance = round(float(row[4]), 2)
                previous_balance = round(new_balance - amount, 2)

                if abs(previous_balance) < 1e-9:
                    cur.execute("DELETE FROM ledger_balances WHERE customer_id = %s", (customer_id,))
                else:
                    cur.execute(
                        """
                        INSERT INTO ledger_balances(customer_id, balance, updated_at)
                        VALUES (%s, %s, NOW())
                        ON CONFLICT (customer_id)
                        DO UPDATE SET balance = EXCLUDED.balance, updated_at = NOW()
                        """,
                        (customer_id, previous_balance),
                    )
                cur.execute("DELETE FROM transactions WHERE id = %s", (tx_id,))
                self._audit(
                    cur,
                    shop_id=shop_id,
                    action="undo_last_transaction",
                    entity_type="transaction",
                    entity_id=str(tx_id),
                    before_state={"customer_name": customer_name, "amount": amount, "new_balance": new_balance},
                    after_state={"customer_name": customer_name, "new_balance": previous_balance},
                )
                return {
                    "customer_name": self._normalize_name(customer_name),
                    "amount": amount,
                    "new_balance": previous_balance,
                }

    def get_pending_ledgers_with_age(self) -> list[dict]:
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT c.name_norm, c.name, lb.balance
                    FROM customers c
                    JOIN ledger_balances lb ON lb.customer_id = c.id
                    WHERE c.shop_id = %s AND lb.balance > 0
                    """,
                    (shop_id,),
                )
                current_rows = cur.fetchall()
                current_balances = {
                    str(r[0]): {"name": str(r[1]), "balance": round(float(r[2]), 2)}
                    for r in current_rows
                }
                if not current_balances:
                    return []

                cur.execute(
                    """
                    SELECT t.created_at, c.name_norm, t.amount
                    FROM transactions t
                    JOIN customers c ON c.id = t.customer_id
                    WHERE t.shop_id = %s
                    ORDER BY t.created_at ASC, t.id ASC
                    """,
                    (shop_id,),
                )
                tx_rows = cur.fetchall()

                state: dict[str, float] = {}
                became_pending_at: dict[str, datetime | None] = {}
                for ts_raw, name_norm, amount_raw in tx_rows:
                    key = str(name_norm)
                    amount = round(float(amount_raw), 2)
                    prev = state.get(key, 0.0)
                    now_balance = round(prev + amount, 2)
                    state[key] = now_balance
                    if prev <= 0 < now_balance:
                        ts_val = ts_raw
                        if isinstance(ts_val, datetime) and ts_val.tzinfo is not None:
                            ts_val = ts_val.replace(tzinfo=None)
                        became_pending_at[key] = ts_val if isinstance(ts_val, datetime) else datetime.now()
                    if now_balance <= 0:
                        became_pending_at[key] = None

                now_dt = datetime.now()
                rows: list[dict] = []
                for name_norm, snapshot in current_balances.items():
                    started = became_pending_at.get(name_norm)
                    days = 0 if started is None else max(0, (now_dt - started).days)
                    rows.append(
                        {
                            "name": snapshot["name"],
                            "balance": snapshot["balance"],
                            "pending_days": days,
                        }
                    )
                rows.sort(key=lambda x: (-float(x["balance"]), str(x["name"])))
                return rows

    def merge_customers(self, source_name: str, target_name: str) -> dict:
        source = self._normalize_name(source_name)
        target = self._normalize_name(target_name)
        if not source or not target:
            return {"ok": False, "reason": "source/target name missing"}
        if source == target:
            return {"ok": False, "reason": "source and target are same"}

        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    "SELECT id, name FROM customers WHERE shop_id = %s AND name_norm = %s",
                    (shop_id, source),
                )
                src_row = cur.fetchone()
                if src_row is None:
                    return {"ok": False, "reason": f"'{source}' not found"}
                source_id = int(src_row[0])

                cur.execute(
                    "SELECT id, name FROM customers WHERE shop_id = %s AND name_norm = %s",
                    (shop_id, target),
                )
                tgt_row = cur.fetchone()
                if tgt_row is None:
                    target_id, _ = self._get_or_create_customer(cur, shop_id, target)
                else:
                    target_id = int(tgt_row[0])

                cur.execute("SELECT balance FROM ledger_balances WHERE customer_id = %s", (source_id,))
                source_bal_row = cur.fetchone()
                source_balance = round(float(source_bal_row[0]), 2) if source_bal_row else 0.0
                cur.execute("SELECT balance FROM ledger_balances WHERE customer_id = %s", (target_id,))
                target_bal_row = cur.fetchone()
                target_balance = round(float(target_bal_row[0]), 2) if target_bal_row else 0.0
                merged_balance = round(source_balance + target_balance, 2)

                if abs(merged_balance) < 1e-9:
                    cur.execute("DELETE FROM ledger_balances WHERE customer_id = %s", (target_id,))
                else:
                    cur.execute(
                        """
                        INSERT INTO ledger_balances(customer_id, balance, updated_at)
                        VALUES (%s, %s, NOW())
                        ON CONFLICT (customer_id)
                        DO UPDATE SET balance = EXCLUDED.balance, updated_at = NOW()
                        """,
                        (target_id, merged_balance),
                    )
                cur.execute("DELETE FROM ledger_balances WHERE customer_id = %s", (source_id,))
                cur.execute("UPDATE transactions SET customer_id = %s WHERE customer_id = %s", (target_id, source_id))
                history_updates = int(cur.rowcount)
                cur.execute("DELETE FROM customers WHERE id = %s", (source_id,))
                self._audit(
                    cur,
                    shop_id=shop_id,
                    action="merge_customers",
                    entity_type="customer",
                    entity_id=str(target_id),
                    before_state={"source": source, "target": target, "target_balance": target_balance},
                    after_state={"target": target, "new_balance": merged_balance},
                    meta={"history_updates": history_updates},
                )
                return {
                    "ok": True,
                    "source": source,
                    "target": target,
                    "history_updates": history_updates,
                    "new_balance": merged_balance,
                }

    def cleanup_noisy_customer_names(self) -> dict:
        with self._connect() as conn:
            with conn.cursor() as cur:
                shop_id = self._get_shop_id(cur)
                cur.execute(
                    """
                    SELECT c.name
                    FROM customers c
                    JOIN ledger_balances lb ON lb.customer_id = c.id
                    WHERE c.shop_id = %s
                    """,
                    (shop_id,),
                )
                names = [self._normalize_name(str(r[0])) for r in cur.fetchall() if r and r[0]]

        mappings: dict[str, str] = {}
        for old_name in names:
            new_name = self._canonicalize_customer_name(old_name)
            if new_name and new_name != old_name:
                mappings[old_name] = new_name

        if not mappings:
            return {"ok": True, "updated": 0, "details": []}

        details: list[dict] = []
        for old_name, new_name in mappings.items():
            result = self.merge_customers(old_name, new_name)
            if result.get("ok"):
                details.append(result)
        return {"ok": True, "updated": len(details), "details": details}


def create_db(filepath: str | None = None):
    backend = os.getenv("DATA_BACKEND", "excel").strip().lower()
    if backend == "postgres":
        return PostgresKhataDB()
    return KhataDB(filepath=filepath)
